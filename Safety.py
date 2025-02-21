import streamlit as st
import pandas as pd
import os
from openpyxl import load_workbook
import matplotlib.pyplot as plt

# Set the page layout to "wide"
st.set_page_config(layout="wide")

st.title("1.Safety")

# Function to write data to Excel
def write_to_excel(date, issue, owner, comment, job_number):
    file_path = "data.xlsx"
    
    # Check if the file exists
    if os.path.exists(file_path):
        # Load the existing workbook
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        # Create a new workbook if the file doesn't exist
        wb = load_workbook("data.xlsx")
        ws = wb.active
        ws.append(["Date", "Issue", "Owner", "Comment", "Job#"])  # Column headers

    # Append new data to the sheet
    ws.append([date, issue, owner, comment, job_number])

    # Save the workbook
    wb.save(file_path)

# Load the Excel file into a DataFrame
def load_excel(file_path):
    try:
        return pd.read_excel(file_path)
    except Exception as e:
        st.error(f"Error reading Excel file: {e}")
        return pd.DataFrame()

# File path for Excel (you can change this to the location you want)
file_path = "data.xlsx"

# Check if the Excel file exists
if not os.path.exists(file_path):
    st.write("No Excel file found. Please create one first.")

# Load the Excel data into a DataFrame
df = load_excel(file_path)

# Initialize the list of issues if it's not already in session_state
if 'issues_list' not in st.session_state:
    st.session_state.issues_list = ['Coating', 'Chip off', 'Damage']  # Initial predefined list

# Create columns for the layout above the displayed data
col1, col2, col3, col4, col5 = st.columns([1, 1, 2, 1, 1])  # Adjust column widths as needed

# Add widgets to each column for input fields
with col1:
    date_input = st.date_input("Select a date")

with col2:
    input1 = st.text_input("Job#")

with col3:
    # Allow the user to select from the predefined issues and add new ones
    input2 = st.selectbox("Select Issue", st.session_state.issues_list)
    
    # Allow adding a custom issue
    custom_issue = st.text_input("Or enter a custom issue (optional)", "")
    
    # If a custom issue is provided, add it to the dropdown list
    if custom_issue:
        if custom_issue not in st.session_state.issues_list:  # Avoid duplicates
            st.session_state.issues_list.append(custom_issue)

with col4:
    input3 = st.text_input("Owner")

with col5:
    input4 = st.text_input("Comment")

# Show the Excel data always
st.write("### Excel Data")
st.dataframe(df, use_container_width=True)

# Generate and display the Pareto Chart for 'Issue' column
def generate_pareto_chart(df):
    if "Issue" in df.columns:
        # Count the frequency of each unique issue
        issue_counts = df["Issue"].value_counts()

        # Calculate the cumulative percentage
        issue_counts_cum = issue_counts.cumsum() / issue_counts.sum() * 100

        # Create a figure for the Pareto chart
        fig, ax1 = plt.subplots(figsize=(10, 6))

        # Plot the bar chart (frequency of each issue)
        ax1.bar(issue_counts.index, issue_counts.values, color='blue', alpha=0.6, label='Frequency')

        # Plot the cumulative percentage line
        ax2 = ax1.twinx()
        ax2.plot(issue_counts.index, issue_counts_cum, color='red', marker='o', label='Cumulative Percentage', linestyle='--')

        # Title and labels
        ax1.set_title("Pareto Chart of Issues")
        ax1.set_xlabel("Issue")
        ax1.set_ylabel("Frequency")
        ax2.set_ylabel("Cumulative Percentage (%)")

        # Display the percentage on the chart
        for i, v in enumerate(issue_counts.values):
            ax1.text(i, v + 0.5, str(v), ha='center', va='bottom')

        for i, v in enumerate(issue_counts_cum):
            ax2.text(i, v + 1, f"{v:.1f}%", ha='center', va='bottom')

        # Show the plot
        st.pyplot(fig)
    else:
        st.write("The 'Issue' column is missing from the dataset.")

# Always display Pareto Chart
generate_pareto_chart(df)

# Check if required fields are filled out
if date_input and input1 and input2 and input3:
    if st.button("Save New Entry and Display Data"):
        # Save new entry to Excel
        write_to_excel(date_input, input1, input2, input3, input4)
        
        # Reload the data after saving it
        df = load_excel(file_path)
        
        # Display the updated Excel data in the app
        st.success("New data has been written to Excel and displayed!")
else:
    if st.button("Save New Entry and Display Data"):
        st.error("Please fill out all the required fields (Date, Job#, Owner, Issue).")
